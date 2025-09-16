
import pandas as pd
import jinja2 as jinja
import pdfkit
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from django.shortcuts import render
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView, View
from django_pandas.io import read_frame
from django.db.models import Sum
from django.db.models import Q
from django.http import HttpResponse
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse
from rest_framework import generics
from rest_framework.response import Response
from docxtpl import DocxTemplate
from nluis_collect.models import FormAnswer, FormField, FormAnswerQuestionnaire
# import dataframe_image as dfi
# Import mimetypes module
import mimetypes
# import os module
import os

from nluis_projects.models import Project, ProjectType
from nluis_ccro.models import Parcel, Party, Allocation
from nluis_spatial.models import SpatialUnit
from nluis_localities.models import Locality
from nluis_reports.models import Report
from nluis_reports.api.serializers import ReportListSerializer
from nluis_projects.models import (Chapter)


# Create your views here.

class ReportListView(generics.ListAPIView):
    serializer_class = ReportListSerializer

    def get_queryset(self):
        return Report.objects.all()


class LandUsesAreaView(View):

    def get(self, request):
        source = {}
        list = []
        try:
            project_type = ProjectType.objects.get(code='VLUS')

            spacial_units = SpatialUnit.objects.values('locality__parent__parent__name', 'locality__name', 'land_use__name').filter(
                project__project_type=project_type).annotate(sqm_area=Sum('sqm')).order_by()

            if 'locality_id' in request.GET and request.GET['locality_id'] is not None:
                _locality_id = request.GET['locality_id']
                spacial_units = spacial_units.filter(locality_id=_locality_id)

            for landUse in spacial_units:
                list.append({
                    'Uses': landUse['land_use__name'],
                    'sqm': landUse['sqm_area'],
                    'Village': landUse['locality__name'],
                    'District': landUse['locality__parent__parent__name'],
                })

            df = pd.DataFrame(list)

            pivot = df.pivot_table(
                index=['Village', 'District'], columns='Uses', values='sqm')

            source = {
                "title": "Land Uses Per km sq",
                "df": pivot.to_html()
            }

        except Exception as e:
            print(str(e))

        if 'format' in request.GET:
            format = request.GET['format']
            return download(self, format, pivot, 'land-use', source)

        return render(request, 'sample.html', context=source)


class FunderView(View):
    def get(self, request):
        source = {}
        list = []
        try:
            project_type = ProjectType.objects.get(code='VLUS')

            funders = Project.objects.values('localites__parent__parent__name', 'localites__name', 'funders__name').filter(
                project_type=project_type).order_by()

            if 'locality_id' in request.GET and request.GET['locality_id'] is not None:
                _locality_id = request.GET['locality_id']
                locality = Locality.objects.get(id=_locality_id)
                funders = funders.prefetch_related('localites').filter(
                    localites__in=locality.getChildrenIds())

            for funder in funders:
                list.append({
                    'Village': funder['localites__name'],
                    # 'District': funder['localites__parent__name'],
                    'District': funder['localites__parent__parent__name'],
                    'Funder': funder['funders__name']
                })

            df = pd.DataFrame(list)

            source = {
                "title": "LIST OF LAND USE PER FUNDERS",
                "df": df.to_html()
            }

        except Exception as e:
            print(str(e))

        if 'format' in request.GET:
            format = request.GET['format']
            return download(self, format, df, 'funders', source)
        print(funders)
        return render(request, 'sample.html', context=source)


class SpecialPartiesView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, **kwargs):
        source = {}
        list = []
        allocations = ''
        try:
            village_id = kwargs['village_id']
            locality = Locality.objects.get(id=village_id)
            parcels = Parcel.objects.filter(Q(locality=locality, stage='printed') | Q(locality=locality,
                                                                                      stage='registered'))
            allocations = Allocation.objects.filter(parcel__in=parcels)

            for allocation in allocations:
                party = allocation.party
                list.append({
                    'party': party.first_name + ' ' + party.middle_name + ' ' + party.last_name,
                    'gender': party.gender,
                    'marital': party.marital,
                    'disability': party.disability
                })

            df = pd.DataFrame(list)

            source = {
                "title": "Special Parties",
                "df": df.to_html()
            }

        except Exception as e:
            print(str(e))

        # return Response({
        #     'results': list
        # })

        if 'format' in request.GET:
            format = request.GET['format']
            return download(self, format, df, 'special-parties', source)
        print(list)
        return render(request, 'sample.html', context=source)


def download(self, format, df, filename, context):
    if format == 'graph':
        numeric_data = df.dtypes.map(pd.api.types.is_numeric_dtype)
        if any(numeric_data):
            df.plot.bar()
            plt.savefig(settings.REPORT_PDF_DIR + '/'+filename)
            fs = FileSystemStorage(settings.REPORT_PDF_DIR+'/')
            response = FileResponse(fs.open(filename +
                                            '.png', 'rb'), content_type='application/force-download')
            response['Content-Disposition'] = 'attachment; filename="' + \
                filename+'.png"'
            return response
        return render(self.request, 'sample.html', context=context)

    if format == 'excel':
        response = HttpResponse(content_type='application/xlsx')
        response['Content-Disposition'] = f'attachment; filename="' + \
            filename+'.xlsx"'
        with pd.ExcelWriter(response) as writer:
            df.to_excel(writer, sheet_name=filename)
        return response

    if format == 'pdf':
        template_loader = jinja.FileSystemLoader(settings.REPORT_PDF_DIR)
        template_env = jinja.Environment(loader=template_loader)

        template = template_env.get_template('sample.html')
        output_text = template.render(context)

        config = pdfkit.configuration()
        pdfkit.from_string(output_text, settings.REPORT_PDF_DIR +
                           '/'+filename+'.pdf', configuration=config, css=settings.REPORT_PDF_DIR+'/sample.css')

        fs = FileSystemStorage(settings.REPORT_PDF_DIR+'/')
        response = FileResponse(fs.open(filename +
                                        '.pdf', 'rb'), content_type='application/force-download')
        response['Content-Disposition'] = 'attachment; filename="' + \
            filename+'.pdf"'
        return response


class LandUseGenerateDocView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):

        try:

            if 'project_id' in request.GET:
                project_id = request.GET['project_id']
                locality_id = request.GET['locality_id']
                project = Project.objects.select_related(
                    'project_type').get(id=project_id)

                locality = Locality.objects.get(id=locality_id)
                ward = Locality.objects.get(id=locality.parent.id)
                hamlet = None
                district = Locality.objects.get(id=ward.parent.id)
                region = Locality.objects.get(id=district.parent.id)

                code = project.project_type.code

                doc = DocxTemplate(code+".docx")
                context = {
                    'kijiji': locality.name,
                    'kata': ward.name,
                    'tarafa': hamlet,
                    'wilaya': district.name,
                    'mkoa': region.name,
                }
                # BASE_DIR = os.path.dirname(
                # os.path.dirname(os.path.abspath(__file__)))
                # doc.render(context)
                # doc.save(BASE_DIR+'/'+locality.name+".docx")

                answers = FormAnswer.objects.filter(project_id=project.id)

                for answer in answers:
                    form_field = FormField.objects.get(id=answer.form_field.id)
                    if form_field.param_code is not None:
                        context.update(
                            {form_field.param_code: answer.response})

                BASE_DIR = os.path.dirname(
                    os.path.dirname(os.path.abspath(__file__)))
                doc.render(context)
                doc.save(BASE_DIR+'/'+locality.name+".docx")

                return Response({'message': code + " was generated succesfully"})
        except Exception as e:
            return Response({'message': str(e)})


class DownloadDocView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):

        try:

            if 'project_id' in request.GET:
                project_id = request.GET['project_id']
                locality_id = request.GET['locality_id']
                project = Project.objects.select_related(
                    'project_type').get(id=project_id)

                locality = Locality.objects.get(id=locality_id)
                code = project.project_type.code

                # Define Django project base directory
                BASE_DIR = os.path.dirname(
                    os.path.dirname(os.path.abspath(__file__)))
                # Define text file name
                filename = locality.name+".docx"
                # Define the full file path
                filepath = BASE_DIR+'/' + filename
                print(filepath)
                # Open the file for reading content
                path = open(filepath, 'rb')
                # Set the mime type
                mime_type, _ = mimetypes.guess_type(filepath)
                # Set the return value of the HttpResponse
                response = HttpResponse(path, content_type=mime_type)
                # Set the HTTP header for sending to browser
                response['Content-Disposition'] = "attachment; filename=%s" % filename
                # Return the response value
                return response
        except Exception as e:
            return Response({'message': str(e)})





# @permission_classes([AllowAny])
def generate_excel(request, project_id, village_name):
    # Create a new Excel workbook and add a worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active

    # Add data to the worksheet
    ws['A1'] = 'Tarehe ya Usajili'
    ws['B1'] = 'Kijiji'
    ws['C1'] = 'Namba ya Cheti cha Ardhi ya Kijiji'
    ws['D1'] = 'Jina la Mmiliki'
    ws['E1'] = 'Namba ya Hakimiliki'
    ws['F1'] = 'Kaskazini'
    ws['G1'] = 'Kusini'
    ws['H1'] = 'Mashariki'
    ws['I1'] = 'Magharibi'
    ws['J1'] = 'Eneo'
    ws['K1'] = 'Matumizi'

    # data = [
    #     ('Alice', 25),
    #     ('Bob', 30),
    #     ('Charlie', 22),
    # ]
    # for row_num, (name, age) in enumerate(data, start=2):
    #     ws.cell(row=row_num, column=1, value=name)
    #     ws.cell(row=row_num, column=2, value=age)

    row_num = 2
    for d in Parcel.objects.filter(uka_namba__isnull=False).filter(project_id=project_id,
                                                                   locality__parent__name__ilike=str(village_name).lower()).order_by(
            'registration_no'):
        ws.cell(row=row_num, column=1, value='')
        ws.cell(row=row_num, column=2, value=d.locality.parent.name.upper())
        ws.cell(row=row_num, column=3, value=d.registration_no.upper())
        ws.cell(row=row_num, column=4, value=str(d.parties()).upper())
        ws.cell(row=row_num, column=5, value=d.uka_namba.upper())
        ws.cell(row=row_num, column=6, value=d.north.upper())
        ws.cell(row=row_num, column=7, value=d.south.upper())
        ws.cell(row=row_num, column=8, value=d.east.upper())
        ws.cell(row=row_num, column=9, value=d.west.upper())
        ws.cell(row=row_num, column=10, value=d.area())
        ws.cell(row=row_num, column=11, value=d.current_use.name.upper())
        row_num += 1

    # Create the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={village_name}.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response
    
